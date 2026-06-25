"""
relatorio_vendas_auto.py
========================
Automação de relatório de vendas em Excel.

Lê um arquivo de dados de vendas (CSV) e gera uma planilha Excel (.xlsx)
profissional com 4 abas formatadas:

    1. Resumo          -> indicadores principais (KPIs) do período
    2. Por Categoria   -> faturamento e quantidade agrupados por categoria
    3. Por Vendedor    -> desempenho de cada vendedor
    4. Dados Brutos    -> a base completa, com filtros e formatação

Se o arquivo de entrada não existir, o script gera automaticamente uma
base de exemplo, para que o relatório possa ser demonstrado na hora.

Uso:
    py relatorio_vendas_auto.py
    py relatorio_vendas_auto.py --entrada vendas.csv --saida relatorio.xlsx

Autor: Bruno Porto
Dependências: pandas, openpyxl  (pip install pandas openpyxl)
"""

import argparse
import random
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# --------------------------------------------------------------------------- #
# Paleta de cores e estilos (centralizados para fácil personalização)
# --------------------------------------------------------------------------- #
COR_CABECALHO = "1F4E78"      # azul escuro
COR_TEXTO_CABECALHO = "FFFFFF"  # branco
COR_FAIXA = "D9E1F2"          # azul bem claro (linhas alternadas)
COR_TITULO = "2E75B6"         # azul médio (títulos das abas)
FORMATO_MOEDA = 'R$ #,##0.00'
FORMATO_INTEIRO = '#,##0'

BORDA_FINA = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# --------------------------------------------------------------------------- #
# 1. Geração de dados de exemplo (caso não haja arquivo de entrada)
# --------------------------------------------------------------------------- #
def gerar_dados_exemplo(caminho: Path, n_registros: int = 300) -> None:
    """Cria um CSV de vendas fictícias para demonstração."""
    random.seed(42)  # resultados reproduzíveis

    categorias = {
        "Eletrônicos": ["Fone Bluetooth", "Smartwatch", "Carregador", "Caixa de Som"],
        "Informática": ["Mouse Gamer", "Teclado Mecânico", "Webcam HD", "SSD 480GB"],
        "Casa": ["Luminária LED", "Organizador", "Cafeteira", "Panela Antiaderente"],
        "Esporte": ["Garrafa Térmica", "Corda de Pular", "Tapete Yoga", "Halteres 5kg"],
    }
    vendedores = ["Ana Souza", "Bruno Lima", "Carla Dias", "Diego Alves", "Eduarda Reis"]
    regioes = ["Sudeste", "Sul", "Nordeste", "Centro-Oeste", "Norte"]

    hoje = datetime.now()
    linhas = []
    for i in range(1, n_registros + 1):
        categoria = random.choice(list(categorias.keys()))
        produto = random.choice(categorias[categoria])
        quantidade = random.randint(1, 8)
        preco_unit = round(random.uniform(25, 850), 2)
        data = hoje - timedelta(days=random.randint(0, 89))
        linhas.append({
            "id_venda": f"V{i:04d}",
            "data": data.strftime("%Y-%m-%d"),
            "produto": produto,
            "categoria": categoria,
            "vendedor": random.choice(vendedores),
            "regiao": random.choice(regioes),
            "quantidade": quantidade,
            "preco_unitario": preco_unit,
            "valor_total": round(quantidade * preco_unit, 2),
        })

    pd.DataFrame(linhas).to_csv(caminho, index=False, encoding="utf-8-sig")
    print(f"  -> Base de exemplo gerada em: {caminho.name} ({n_registros} registros)")


# --------------------------------------------------------------------------- #
# 2. Leitura e preparação dos dados
# --------------------------------------------------------------------------- #
def carregar_dados(caminho: Path) -> pd.DataFrame:
    """Lê o CSV de vendas e garante os tipos corretos."""
    df = pd.read_csv(caminho, encoding="utf-8-sig")
    df["data"] = pd.to_datetime(df["data"])

    # Recalcula valor_total por segurança (caso a coluna venha ausente/errada)
    df["valor_total"] = (df["quantidade"] * df["preco_unitario"]).round(2)
    return df


# --------------------------------------------------------------------------- #
# 3. Cálculo das visões agregadas
# --------------------------------------------------------------------------- #
def calcular_resumo(df: pd.DataFrame) -> list[tuple[str, object, str]]:
    """Retorna uma lista de (indicador, valor, formato) para a aba Resumo."""
    faturamento = df["valor_total"].sum()
    n_pedidos = len(df)
    itens_vendidos = int(df["quantidade"].sum())
    ticket_medio = faturamento / n_pedidos if n_pedidos else 0
    top_cat = df.groupby("categoria")["valor_total"].sum().idxmax()
    top_vend = df.groupby("vendedor")["valor_total"].sum().idxmax()
    periodo = f"{df['data'].min():%d/%m/%Y} a {df['data'].max():%d/%m/%Y}"

    return [
        ("Período analisado", periodo, "texto"),
        ("Faturamento total", faturamento, "moeda"),
        ("Número de pedidos", n_pedidos, "inteiro"),
        ("Itens vendidos", itens_vendidos, "inteiro"),
        ("Ticket médio", ticket_medio, "moeda"),
        ("Categoria destaque", top_cat, "texto"),
        ("Vendedor destaque", top_vend, "texto"),
    ]


def agrupar_por(df: pd.DataFrame, coluna: str) -> pd.DataFrame:
    """Agrupa o faturamento e a quantidade por uma coluna (categoria/vendedor)."""
    agrupado = (
        df.groupby(coluna)
        .agg(
            faturamento=("valor_total", "sum"),
            itens_vendidos=("quantidade", "sum"),
            n_pedidos=("id_venda", "count"),
        )
        .reset_index()
        .sort_values("faturamento", ascending=False)
    )
    agrupado["ticket_medio"] = (agrupado["faturamento"] / agrupado["n_pedidos"]).round(2)
    agrupado["% do total"] = (
        agrupado["faturamento"] / agrupado["faturamento"].sum() * 100
    ).round(1)
    return agrupado


# --------------------------------------------------------------------------- #
# 4. Funções de formatação do Excel (openpyxl)
# --------------------------------------------------------------------------- #
def estilizar_cabecalho(ws, linha: int, n_colunas: int) -> None:
    """Aplica o estilo padrão de cabeçalho a uma linha."""
    fill = PatternFill("solid", fgColor=COR_CABECALHO)
    fonte = Font(bold=True, color=COR_TEXTO_CABECALHO, size=11)
    for col in range(1, n_colunas + 1):
        celula = ws.cell(row=linha, column=col)
        celula.fill = fill
        celula.font = fonte
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = BORDA_FINA


def ajustar_larguras(ws) -> None:
    """Ajusta a largura das colunas com base no maior conteúdo.

    Percorre por índice de coluna e ignora células mescladas (o título),
    que não possuem um valor próprio nem letra de coluna.
    """
    for col_idx in range(1, ws.max_column + 1):
        letra = get_column_letter(col_idx)
        comprimentos = []
        for row_idx in range(1, ws.max_row + 1):
            celula = ws.cell(row=row_idx, column=col_idx)
            # Pula células mescladas (título) para não contar o texto duas vezes
            if celula.__class__.__name__ == "MergedCell":
                continue
            if celula.value is not None:
                comprimentos.append(len(str(celula.value)))
        largura = max(comprimentos) + 3 if comprimentos else 12
        ws.column_dimensions[letra].width = min(largura, 40)


def escrever_titulo(ws, texto: str, n_colunas: int) -> None:
    """Escreve um título grande mesclado no topo da aba."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_colunas)
    celula = ws.cell(row=1, column=1, value=texto)
    celula.font = Font(bold=True, size=14, color=COR_TITULO)
    celula.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


# --------------------------------------------------------------------------- #
# 5. Construção de cada aba
# --------------------------------------------------------------------------- #
def montar_aba_resumo(ws, resumo: list) -> None:
    escrever_titulo(ws, "Resumo Geral de Vendas", 2)
    ws.cell(row=2, column=1, value="Indicador")
    ws.cell(row=2, column=2, value="Valor")
    estilizar_cabecalho(ws, 2, 2)

    for i, (nome, valor, formato) in enumerate(resumo, start=3):
        c_nome = ws.cell(row=i, column=1, value=nome)
        c_valor = ws.cell(row=i, column=2, value=valor)
        c_nome.font = Font(bold=True)
        c_nome.border = BORDA_FINA
        c_valor.border = BORDA_FINA
        c_valor.alignment = Alignment(horizontal="right")
        if formato == "moeda":
            c_valor.number_format = FORMATO_MOEDA
        elif formato == "inteiro":
            c_valor.number_format = FORMATO_INTEIRO
        if i % 2 == 0:
            for col in (1, 2):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=COR_FAIXA)

    ajustar_larguras(ws)
    ws.column_dimensions["B"].width = 28


def montar_aba_dataframe(ws, df: pd.DataFrame, titulo: str,
                         colunas_moeda=(), colunas_percent=()) -> None:
    """Escreve um DataFrame numa aba com cabeçalho, faixas e formatação."""
    escrever_titulo(ws, titulo, len(df.columns))

    linha_cabecalho = 2
    for r_idx, linha in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, valor in enumerate(linha, start=1):
            ws.cell(row=linha_cabecalho + r_idx, column=c_idx, value=valor)

    estilizar_cabecalho(ws, linha_cabecalho, len(df.columns))

    # Formatação das linhas de dados
    primeira_dado = linha_cabecalho + 1
    ultima_dado = linha_cabecalho + len(df)
    nomes_colunas = list(df.columns)
    for r in range(primeira_dado, ultima_dado + 1):
        faixa = (r - primeira_dado) % 2 == 1
        for c in range(1, len(nomes_colunas) + 1):
            celula = ws.cell(row=r, column=c)
            celula.border = BORDA_FINA
            if faixa:
                celula.fill = PatternFill("solid", fgColor=COR_FAIXA)
            nome_col = nomes_colunas[c - 1]
            if nome_col in colunas_moeda:
                celula.number_format = FORMATO_MOEDA
            elif nome_col in colunas_percent:
                celula.number_format = '0.0"%"'

    # Congela o cabeçalho e liga o filtro
    ws.freeze_panes = ws.cell(row=primeira_dado, column=1)
    ws.auto_filter.ref = (
        f"A{linha_cabecalho}:"
        f"{get_column_letter(len(df.columns))}{ultima_dado}"
    )
    ajustar_larguras(ws)


# --------------------------------------------------------------------------- #
# 6. Orquestração geral
# --------------------------------------------------------------------------- #
def gerar_relatorio(entrada: Path, saida: Path) -> None:
    print("Gerando relatório de vendas...")

    if not entrada.exists():
        print(f"  Arquivo '{entrada.name}' não encontrado.")
        gerar_dados_exemplo(entrada)

    df = carregar_dados(entrada)
    print(f"  -> {len(df)} registros carregados de {entrada.name}")

    resumo = calcular_resumo(df)
    por_categoria = agrupar_por(df, "categoria")
    por_vendedor = agrupar_por(df, "vendedor")

    # Prepara a aba de dados brutos (data formatada como texto dd/mm/aaaa)
    dados_brutos = df.copy()
    dados_brutos["data"] = dados_brutos["data"].dt.strftime("%d/%m/%Y")

    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        # Cria 4 abas vazias e depois preenche via openpyxl para ter controle total
        for nome_aba in ["Resumo", "Por Categoria", "Por Vendedor", "Dados Brutos"]:
            writer.book.create_sheet(nome_aba)
        # Remove a aba padrão criada automaticamente, se existir
        if "Sheet" in writer.book.sheetnames:
            del writer.book["Sheet"]

        montar_aba_resumo(writer.book["Resumo"], resumo)
        montar_aba_dataframe(
            writer.book["Por Categoria"], por_categoria,
            "Faturamento por Categoria",
            colunas_moeda=("faturamento", "ticket_medio"),
            colunas_percent=("% do total",),
        )
        montar_aba_dataframe(
            writer.book["Por Vendedor"], por_vendedor,
            "Desempenho por Vendedor",
            colunas_moeda=("faturamento", "ticket_medio"),
            colunas_percent=("% do total",),
        )
        montar_aba_dataframe(
            writer.book["Dados Brutos"], dados_brutos,
            "Dados Brutos das Vendas",
            colunas_moeda=("preco_unitario", "valor_total"),
        )

    print(f"  -> Relatório salvo em: {saida.name}")
    print("Concluído! Abra o arquivo no Excel para conferir as 4 abas.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera relatório de vendas em Excel.")
    parser.add_argument("--entrada", default="vendas.csv",
                        help="CSV de entrada (padrão: vendas.csv)")
    parser.add_argument("--saida", default="relatorio_vendas.xlsx",
                        help="Excel de saída (padrão: relatorio_vendas.xlsx)")
    args = parser.parse_args()

    pasta = Path(__file__).parent
    entrada = pasta / args.entrada
    saida = pasta / args.saida
    gerar_relatorio(entrada, saida)


if __name__ == "__main__":
    main()
